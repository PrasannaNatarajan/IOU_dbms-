# from xlwt import *
# from xlrd import *
# from xlutils.copy import *
from openpyxl import *
import time

# workbook = xlrd.open_workbook('Book1.xls')
# sheet = workbook.sheet_by_index(0)

# data = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
#
# workbook = xlwt.Workbook()
# sheet = workbook.add_sheet('test')
#
# for index, value in enumerate(data):
#     sheet.write(0, index, value)
#
# workbook.save('output.xls')
with open('tabs.txt', 'r') as f:
    LoL = [x.strip().split('\t') for x in f]
# print (LoL)

# AStuff = {'FC4', 'FD'}

wb = load_workbook(filename='outFile.xlsx')
# ws = wb.get_sheet_by_name('Birds details')


sheet_ranges = wb['Birds details']
sheet2_ranges = wb['Feed']
# print(sheet_ranges.cell(row=4, column=3).value)
that = 159
now = that
touchdown = 1
dateSave = LoL[13][2]
reasonsMed = LoL[13][0]
remarks = LoL[13][1]

for i in range(4, 44):
    # print(str(sheet_ranges.cell(row=i, column=1).value)[:10] + " " + dateSave[6:10]+"-"+dateSave[3:5]+"-"+dateSave[0:2])
    if str(sheet_ranges.cell(row=i, column=1).value)[:10] == dateSave[6:10]+"-"+dateSave[3:5]+"-"+dateSave[0:2]:
        touchdown = i
        # print(touchdown)
        break


ind = 3
# first identify row, then this
for i in range(1, 13):
    for j in range(now, now + 12):
        # print str(int(LoL[i][j-now])) + " " + str(touchdown) + " " + str(j)
        sheet_ranges.cell(row=touchdown, column=j).value = int(LoL[i][j - now])
        sheet2_ranges.cell(row=touchdown, column=ind + j - now).value = float(LoL[i + 13][j-now])
    now = now + 12 + 1
    ind += 13
    # that = now + 12
sheet_ranges.cell(row=touchdown, column=471).value = reasonsMed[0:-1]
sheet_ranges.cell(row=touchdown, column=472).value = remarks[0:-1]


sheet_ranges.cell(row=touchdown, column=473).value = str(time.strftime("%d-%B-%y | %H:%M:%S"))
sheet_ranges.cell(row=touchdown, column=1).value = LoL[13][2]

wb.save('newDetails.xlsx')
# wb = load_workbook(filename='outFile.xlsx')
# sheet2_ranges = wb['Feed']
# that = 159
# now = that
# for i in range(1, 13):
#     for j in range(now, now + 12):
#         # print str(int(LoL[i][j-now])) + " " + str(touchdown) + " " + str(j)
#         # sheet_ranges.cell(row=touchdown, column=j).value = int(LoL[i][j - now])
#         sheet2_ranges.cell(row=touchdown, column=ind + j - now).value = float(LoL[i + 13][j-now])
#     now = now + 12 + 1
#     ind += 13
#

# at 473, put cdx-wala date

#
# print wb.get_sheet_names()
# print ws['F1']

# f.close()
#
# readBook = open_workbook('Book1.xls')
# readSheet = readBook.sheet_by_index(0)
# x = int(LoL[0][0])
# data = readSheet.cell_value(0, 0)


# rb = open_workbook('outFile.xlsx', formatting_info=True)
# rb = open_workbook('outFile.xlsx')
# r_sheet = rb.sheet_by_index(3) # read only copy to introspect the file
# wb = copy(rb) # a writable copy (I can't read values out of this, only write to it)
#
# w_sheet = wb.get_sheet(3) # the sheet to write to within the writable copy

# x = 3
# m = r_sheet.cell_value(x, 1)
# if int(m) == 3:
#     w_sheet.write(x, ord('s')-ord('a'), "colombo")
# x = 4
# if int(r_sheet.cell_value(x, 1)) == 3:
#     w_sheet.write(x, ord('s')-ord('a'), "colombo")
# x = 5
# if int(r_sheet.cell_value(x, 1)) == 3:
#     w_sheet.write(x, ord('s')-ord('a'), "colombo")
# x = 6
# if int(r_sheet.cell_value(x, 1)) == 3:
#     w_sheet.write(x, ord('s')-ord('a'), "colombo")
#
# x = 1
# print r_sheet.cell_value(x, 1)
# x = 2
# print r_sheet.cell_value(x, 1)
# x = 3
# print r_sheet.cell_value(x, 1)
# x = 4
# print r_sheet.cell_value(x, 1)
# x = 5
# print r_sheet.cell_value(x, 1)
# x = 6
# print r_sheet.cell_value(x, 1)





# w_sheet.write(1, 3, "kozhakattai")
# w_sheet.write(13, 13, "kadubu")
# wb.save('outFile.xlsx')

# print data
